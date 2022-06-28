from openpyxl import Workbook
import re
import json

class Comment:
    #경로 넘겨받기
    def __init__(self, filePath):
        self.filePath = filePath
    
    # a_json = open('result/naver_news/news_{search}_naver_{start_date}_{end_date}__202206 copy.json', 'r', encoding='utf-8')
    def read_json(self):
        a_json = open(self.filePath, 'r', encoding='utf-8')
        a_dict = json.load(a_json)
        return a_dict

    #이모티콘 제거
    def remove_emoji(comment):
        emoji_pattern = re.compile("["
                                u"\U0001F600-\U0001F64F"
                                u"\U0001F300-\U0001F5FF"
                                u"\U0001F680-\U0001F6FF"
                                u"\U0001F1E0-\U0001F1FF"
                                "]+", flags=re.UNICODE)
        return emoji_pattern.sub(r'', comment)

    wb = Workbook()
    ws = wb.active
    ws.title = "crawling_data"

    i = 1
    for date in read_json().keys():
        for url in read_json()[date].keys():
            for comment in read_json()[date][url]["comments"]:
                ws.cell(i, 1, date)
                ws.cell(i, 2, remove_emoji(comment))
                i += 1

    wb.save('coment/crawling_data.xlsx')